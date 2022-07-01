import time

import openpyxl
from selenium import webdriver
import constants
from bs4 import BeautifulSoup

class backEnd:
    """ This class contains the backend code of the method."""

    def __init__(self, file, sheet, address, column_to_extract, column_to_insert):
        """ This method initializes the instance variables.
        It is used to link the frontend inputs to the backend."""
        self.filePath = file
        self.sheetName = sheet
        self.websites = address
        self.column_to_append = column_to_insert
        self.column_to_extract = column_to_extract
        #Dictionary with key as the name of item and value as the price.
        self.names = {}
        self.items = []
        self.prices = []
        self.all_website_items = []


    def extract_from_excel(self):
        """ This method extracts data from the given excel file."""
        #Opens the Excel File
        self.ps = openpyxl.load_workbook(self.filePath)
        #Opens the Sheet in the excel file
        self.sheet = self.ps[self.sheetName]
        #Loops over each row in the column to parse to get the data from it.
        #Starts looping from row 2.
        for row in range(2, self.sheet.max_row):
            # each row in the spreadsheet represents information for a particular item.
            item = self.sheet[self.column_to_extract + str(row)].value
            if item != None:
                #It adds that item to self.names dictionary with value being empty string right now. Price is later
                #added to this.
                self.names[item] = ""
        print(self.names)

    def parse_website(self, website):
        #IMPORTANT!!: Need to find a way to find location of chromedriver
        self.driver = webdriver.Chrome(constants.chromedriver_location)
        self.driver.get(website)
        time.sleep(5)
        self.html = self.driver.page_source
        self.soup = BeautifulSoup(self.html, "html.parser")


    def put_in_excel(self):
        for row in range(1, self.sheet.max_row+2):
            # each row in the spreadsheet represents information for a particular item.
            item = self.sheet[self.column_to_extract + str(row)].value
            #It adds that item to self.names dictionary with value being empty string right now. Price is later
            #added to this.
            if self.names.__contains__(item):
                price = self.names[item]
                cell = self.column_to_append+str(row)
                self.sheet[cell] = price
                self.ps.save(self.filePath)