import backEnd
from openpyxl.styles import PatternFill


#CHANGES: ADD NEW ITEMS TO EXCEL SHEET SOMEHOW


class mcdonald(backEnd.backEnd):
    """ This class defines the data extraction process for mcdonalds."""

    def __init__(self, file, sheet, address, column_to_extract, column_to_insert):
        """ Creates an instance of mcdonald."""
        super().__init__(file, sheet, address, column_to_extract, column_to_insert)


    def extract_price(self, item_name):
        """ This checks a particular web link for an item name. It that item is found on that webpage, it stores its
        price in a dictionary corresponding to the item name. If it is not found, (and has not been found before on any other
        webpage, it puts its value as 'Not Available' in self.names. Else, it does nothing."""
        #Finds whether or not that element is there in self.soup
        self.elem = self.soup.find("div", text = item_name)
        #print(item_name + "has self.elem")
        #print(self.elem)
        #If it is found, it does the following:
        if self.elem:
            #Adds its price corresponding to its name in self.names
            self.names[item_name] = self.elem.parent.parent.find("span",{"class":"price pr-1"}).text
            #Adds item name to the list of items whose prices have already been found.
            self.items.append(item_name)
        #If not found, and not found anywhere else, it stores its value as not available.
        elif(not self.items.__contains__(item_name)):
            self.names[item_name] = "Not Available"
        # (item_name+":"+self.names[item_name])

    def extract_price_of_new(self, item_name):
        self.elem = self.soup.find("div", text = item_name)
        #If it is found, it does the following:
        if self.elem:
            #Adds its price corresponding to its name in self.names
            return self.elem.parent.parent.find("span",{"class":"price pr-1"}).text
            #Adds item name to the list of items whose prices have already been found.
        self.driver.quit()

    def get_all_web_items(self):
        """ This lists down all the elements from mcdonalds and stores it in element self.all_website_items"""
        #Scaps website to find all items.
        elem = self.soup.find("div", id = "home-page-wrapper")
        #all_items is a list with all the items' html code format.
        all_items = elem.findAll('div', {"class": "item-title"})
        #self.all_website_items thus is a list of all items on a menu.
        for item in all_items:
            if not self.all_website_items.__contains__(item):
                #To the field all_website_names, this adds an item if it is not already present.
                to_add = item.text
                #if item.text.__contains__("\t"):
                #to_add = item.text[:len(item.text)-1]
                self.all_website_items.append(to_add)
        #print(self.all_website_items)

    def extract_all_prices(self):
        """ This extracts all prices from all the various links used to extract prices from mcdonalds."""
        #Loops over each website given in the list of websites to access
        for section in self.websites:
            #Accesses and opens those websites
            self.parse_website(section)
            #Gets all the items in a list.
            self.get_all_web_items()
            #print(self.all_website_items)
            #For items in the given list of items it extracts their prices.
            for item in self.names:
                self.extract_price(item)
        print(self.all_website_items)
        #def put_in_excel(self, data):
        """ Puts the extracted prices into the excel sheet."""
        #For each item in the dictionary with items and prices
        #for item in self.names:
        #Puts the price into a variable self.prices, which is a list of all the prices.
        #self.prices.append(self.names[item])
        #Calls the superclass put_in_excel to put the prices into the excel sheet
        #super().put_in_excel()


    def handle_new_items(self):
        starting_row = self.sheet.max_row+1
        for new in self.new_items:
            cell_of_items = self.column_to_extract+str(starting_row)
            self.sheet[cell_of_items] = new
            price = self.extract_price_of_new(new)
            cell_of_price = self.column_to_append+str(starting_row)
            self.sheet[cell_of_price] = price
            yellow = "00FFFF00"
            highlight = PatternFill(start_color=yellow,
                                    end_color=yellow,
                                    fill_type='solid')
            self.sheet[cell_of_items].fill = highlight
            self.sheet[cell_of_price].fill = highlight
            self.ps.save(self.filePath)
            starting_row +=1


    def check_if_new(self):
        """Checks if there are any new items. Returns True if there are, False if there are not."""
        for item in self.all_website_items:
            #if item == "'McSpicy Premium Veg Burger'" or "McSpicy Premium Chicken Burger":
            #print("FUCK OFF")
            #print("doing "+item)
            if self.names.__contains__(item):
                #print(item+"is contained in name")
                self.all_website_items.remove(item)
        #print(self.all_website_items)
        self.new_items = {}
        for new_item in self.all_website_items:
            self.new_items[new_item] = ""
        if len(self.new_items) != 0:
            return True
        else:
            return False